import csv
import datetime
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
except ImportError:
    openpyxl = None

from model.report import ReportModel

class ReportController:
    def __init__(self, view, backend="sqlite", mysql_config=None):
        self.view = view
        self.model = ReportModel(backend=backend, mysql_config=mysql_config)
        self.current_position_filter = None
        self.current_month_year = None

    def load_positions(self):
        try:
            positions = self.model.get_positions()
        except RuntimeError as e:
            self.view.show_message("Lỗi", str(e), "error")
            positions = []
        if not positions:
            self.view.show_message("Thông báo", "Không lấy được danh sách chức vụ.", "warning")
        self.view.set_positions(positions)

    def load_seniority(self, position=None):
        self.current_position_filter = position
        try:
            rows = self.model.get_seniority(position)
        except RuntimeError as e:
            self.view.show_message("Lỗi", str(e), "error")
            rows = []
        if not rows:
            self.view.show_message("Thông báo", "Không có dữ liệu thâm niên.", "warning")
        self.view.render_seniority(rows)

    def load_revenue(self, month_year):
        self.current_month_year = month_year
        try:
            month, year = month_year.split("/")
            rows = self.model.get_revenue_by_month(month, year)
            if not rows:
                count = self.model.revenue_exists(month, year)
                if count > 0:
                    self.view.show_message("Lỗi", "Có dữ liệu chi tiết nhưng nhóm doanh thu trả về rỗng (kiểm tra tên cột hoặc truy vấn).", "error")
                else:
                    self.view.show_message("Thông báo", "Không có dữ liệu doanh thu tháng đã chọn.", "warning")
            total = self.model.sum_revenue(rows)
            self.view.render_revenue(rows, total)
        except RuntimeError as e:
            self.view.show_message("Lỗi", str(e), "error")
            self.view.render_revenue([], 0)
        except ValueError:
            self.view.show_message("Lỗi", "Định dạng tháng/năm không hợp lệ (MM/yyyy).", "error")
            self.view.render_revenue([], 0)

    def export_seniority(self, rows):
        self._export_generic(
            rows,
            base_name="bao_cao_tham_nien",
            report_type="seniority"
        )

    def export_revenue(self, rows):
        self._export_generic(
            rows,
            base_name="bao_cao_doanh_thu",
            report_type="revenue"
        )

    def _export_generic(self, rows, base_name, report_type):
        if not rows:
            self.view.show_message("Thông báo", "Không có dữ liệu để xuất.", "info")
            return

        # Mapping khóa -> tiêu đề hiển thị tiếng Việt (có dấu, khoảng trắng)
        SENIORITY_HEADER_MAP = {
            "ma_nv": "Mã nhân viên",
            "ho_va_ten": "Họ và tên",
            "chuc_vu": "Chức vụ",
            "ngay_vao_lam": "Ngày vào làm",
            "tham_nien": "Thâm niên",
            "nhom_tham_nien": "Nhóm thâm niên"
        }
        REVENUE_HEADER_MAP = {
            "ma_thuoc": "Mã thuốc",
            "ten_thuoc": "Tên thuốc",
            "so_luong_ban": "Số lượng bán",
            "don_gia": "Đơn giá",
            "tong_doanh_thu": "Tổng doanh thu"
        }

        header_map = SENIORITY_HEADER_MAP if report_type == "seniority" else REVENUE_HEADER_MAP
        # Thứ tự cột theo mapping, chỉ giữ các khóa thực sự có trong dữ liệu
        data_keys = [k for k in header_map.keys() if k in rows[0].keys()]
        display_headers = [header_map[k] for k in data_keys]

        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename_xlsx = f"{base_name}_{timestamp}.xlsx"
        filename_csv = f"{base_name}_{timestamp}.csv"

        if not openpyxl:
            with open(filename_csv, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow([" - BÁO CÁO", base_name, timestamp])
                writer.writerow(display_headers)
                for r in rows:
                    writer.writerow([r.get(k, "") for k in data_keys])
            self.view.show_message("Xuất file", f"Đã xuất (CSV): {filename_csv}", "info")
            return

        wb = openpyxl.Workbook()
        ws = wb.active

        masthead_lines = [
            "Nhà thuốc The Blue",
            "Mã số thuế: 012345",
            "Địa chỉ: Số 10, Trần Phú, Hà Đông, Hà Nội",
            "Điện thoại: 0987654321",
            "Tài khoản Ngân hàng: 0382117403 - MBBank"
        ]
        title_map = {
            "seniority": "BÁO CÁO THÂM NIÊN NHÂN VIÊN",
            "revenue": "BÁO CÁO DOANH THU THEO THÁNG"
        }
        title = title_map.get(report_type, "BÁO CÁO")
        numbering = "Mẫu số: TTTTT0101    Số: 0000"

        col_count = len(display_headers)
        r = 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=col_count//2 or 1)
        ws.merge_cells(start_row=r, start_column=(col_count//2 or 1) + 1, end_row=r, end_column=col_count)
        ws.cell(row=r, column=1).value = masthead_lines[0]
        ws.cell(row=r, column=1).font = Font(bold=True)
        ws.cell(row=r, column=(col_count//2 or 1) + 1).value = title
        ws.cell(row=r, column=(col_count//2 or 1) + 1).font = Font(bold=True, size=14)
        ws.cell(row=r, column=(col_count//2 or 1) + 1).alignment = Alignment(horizontal="center")
        r += 1

        for line in masthead_lines[1:]:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=col_count//2 or 1)
            ws.cell(row=r, column=1).value = line
            r += 1

        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=col_count)
        ws.cell(row=r, column=1).value = numbering
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="right")
        r += 1

        filter_info = ""
        if report_type == "seniority":
            filter_info = f"Chức vụ lọc: {self.current_position_filter if self.current_position_filter else '(Tất cả)'}"
        elif report_type == "revenue":
            filter_info = f"Tháng/Năm: {self.current_month_year}"
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=col_count)
        ws.cell(row=r, column=1).value = filter_info
        r += 1

        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=col_count)
        ws.cell(row=r, column=1).value = f"Thời gian xuất: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        r += 2

        for c, h in enumerate(display_headers, start=1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        r += 1

        for row in rows:
            for c, k in enumerate(data_keys, start=1):
                ws.cell(row=r, column=c, value=row.get(k, ""))
            r += 1
        # Thêm dòng tổng cho báo cáo doanh thu
        if report_type == "revenue":
            try:
                idx_total = data_keys.index("tong_doanh_thu") + 1
                idx_label = data_keys.index("don_gia") + 1
            except ValueError:
                idx_total = len(data_keys)
                idx_label = max(1, idx_total - 1)
            total_sum = sum(row.get("tong_doanh_thu", 0) for row in rows)
            ws.cell(row=r, column=idx_label, value="Tổng cộng").font = Font(bold=True)
            ws.cell(row=r, column=idx_total, value=total_sum).font = Font(bold=True)
            ws.cell(row=r, column=idx_label).alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=idx_total).alignment = Alignment(horizontal="center")
            r += 1

        for c, h in enumerate(display_headers, start=1):
            max_len = len(h)
            for row in rows:
                val = str(row.get(data_keys[c-1], ""))
                if len(val) > max_len:
                    max_len = len(val)
            # Nếu là dòng tổng cộng, cập nhật độ dài nếu cần
            if report_type == "revenue" and h == "Tổng doanh thu":
                tlen = len(str(total_sum))
                if tlen > max_len:
                    max_len = tlen
            ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = max_len + 2

        wb.save(filename_xlsx)
        self.view.show_message("Xuất file", f"Đã xuất: {filename_xlsx}", "info")

    def close(self):
        pass
